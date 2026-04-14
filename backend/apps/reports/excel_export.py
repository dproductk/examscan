"""
Excel export utility using openpyxl.
Generates a workbook with two sheets:
  1. Detail — one row per student per subject
  2. Summary — one row per student with grand total
"""
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_excel_report(students):
    """
    Generate an Excel report from a list of student dicts.
    Each student dict has: roll_number, department, semester, academic_year,
                           subjects: [{subject_code, subject_name, total_marks, max_marks, evaluated_on}],
                           grand_total
    Returns an HttpResponse with the Excel file.
    """
    wb = Workbook()

    # ── Styles ────────────────────────────────────
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1C1D22', end_color='1C1D22', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0'),
    )
    data_align = Alignment(horizontal='center', vertical='center')
    alt_fill = PatternFill(start_color='F7F8FA', end_color='F7F8FA', fill_type='solid')

    # ── Detail Sheet ──────────────────────────────
    ws_detail = wb.active
    ws_detail.title = 'Detail'

    detail_headers = [
        'Roll Number', 'Department', 'Semester', 'Academic Year',
        'Subject Code', 'Subject Name', 'Total Marks', 'Max Marks', 'Evaluated On',
    ]
    for col_idx, header in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row_idx = 2
    for student in students:
        for subj in student.get('subjects', []):
            row_data = [
                student.get('roll_number', ''),
                student.get('department', ''),
                student.get('semester', ''),
                student.get('academic_year', ''),
                subj.get('subject_code', ''),
                subj.get('subject_name', ''),
                subj.get('total_marks', 0),
                subj.get('max_marks', 0),
                subj.get('evaluated_on', ''),
            ]
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_detail.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = data_align
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
            row_idx += 1

    # Auto-width
    for col in ws_detail.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_detail.column_dimensions[col_letter].width = max_len + 4

    # ── Summary Sheet ─────────────────────────────
    ws_summary = wb.create_sheet('Summary')

    summary_headers = [
        'Roll Number', 'Department', 'Semester', 'Academic Year',
        'Subjects Evaluated', 'Grand Total',
    ]
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, student in enumerate(students, 2):
        row_data = [
            student.get('roll_number', ''),
            student.get('department', ''),
            student.get('semester', ''),
            student.get('academic_year', ''),
            len(student.get('subjects', [])),
            student.get('grand_total', 0),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = data_align
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col in ws_summary.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_summary.column_dimensions[col_letter].width = max_len + 4

    # ── Response ──────────────────────────────────
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="evaluation_results.xlsx"'
    wb.save(response)
    return response
